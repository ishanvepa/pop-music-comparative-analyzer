import openpyxl as xlsxreader
import math
import collections
import operator
from billboard_scraper import get_tonic

def get_tonic(spreadsheet):
    tonic_c = 0
    tonic_dflat = 0
    tonic_d = 0 
    tonic_eflat = 0
    tonic_e = 0
    tonic_f = 0
    tonic_fsharp = 0
    tonic_g = 0
    tonic_aflat = 0
    tonic_a = 0
    tonic_bflat = 0
    tonic_b = 0
    tonics = {}

    for j in range(11):
        cell = spreadsheet.cell(row = 7, column = j+1)
        if(cell.value == "C" or cell.value == "B#"):
            tonic_c+=1
            tonics[cell.value] = tonic_c
        if(cell.value == "Db" or cell.value == "C#"):
            tonic_dflat+=1
            tonics[cell.value] = tonic_dflat
        if(cell.value == "D"):
            tonic_d+=1
            tonics[cell.value] = tonic_d
        if(cell.value == "Eb" or cell.value == "D#"):
            tonic_eflat+=1
            tonics[cell.value] = tonic_eflat
        if(cell.value == "E" or cell.value == "Fb"):
            tonic_e+=1
            tonics[cell.value] = tonic_e
        if(cell.value == "F" or cell.value == "E#"):
            tonic_f+=1
            tonics[cell.value] = tonic_f
        if(cell.value == "F#" or cell.value == "Gb"):
            tonic_fsharp+=1
            tonics[cell.value] = tonic_fsharp
        if(cell.value == "G"):
            tonic_g+=1
            tonics[cell.value] = tonic_g
        if(cell.value == "Ab" or cell.value == "G#"):
            tonic_aflat+=1
            tonics[cell.value] = tonic_aflat
        if(cell.value == "A"):
            tonic_a+=1
            tonics[cell.value] = tonic_a
        if(cell.value == "Bb" or cell.value == "A#"):
            tonic_bflat+=1
            tonics[cell.value] = tonic_bflat
        if(cell.value == "B" or cell.value == "Cb"):
            tonic_b+=1
            tonics[cell.value] = tonic_b
    return tonics

def get_key_quality(spreadsheet):
    major = 0
    minor = 0
    qualities = {}
    for j in range(11):
        cell = spreadsheet.cell(row = 8, column = j+1)
        if(cell.value == "Major"):
            major+=1
            qualities[cell.value] = major
        if(cell.value == "Minor"):
            minor+=1
            qualities[cell.value] = minor
    return qualities

def get_key_whole(spreadsheet):
    cmajor = 0
    dflatmajor = 0
    dmajor = 0
    eflatmajor = 0
    emajor = 0
    fmajor = 0
    fsharpmajor = 0
    gmajor = 0
    aflatmajor = 0
    amajor = 0
    bflatmajor = 0
    bmajor = 0
    
    cminor = 0
    dflatminor = 0
    dminor = 0
    eflatminor = 0
    eminor = 0
    fminor = 0
    fsharpminor = 0
    gminor = 0
    aflatminor = 0
    aminor = 0
    bflatminor = 0
    bminor = 0
    keys = {}
    
    for j in range(10):
        cell_tonic = spreadsheet.cell(row = 7, column = j+2)
        cell_quality = spreadsheet.cell(row = 8, column = j+2)
        full_key = str(cell_tonic.value) + " " + str(cell_quality.value)
        #major keys
        if str(full_key) not in keys:                
                keys[str(full_key)] = 0
        if(str(full_key).lower() == "C Major".lower()):
            cmajor += 1
            keys[str(full_key)] = cmajor
        if(str(full_key).lower() == "Db Major".lower()):
            dflatmajor += 1
            keys[str(full_key)] = dflatmajor
        if(str(full_key).lower() == "D Major".lower()):
            dmajor += 1
            keys[str(full_key)] = dmajor
        if(str(full_key).lower() == "Eb Major".lower()):
            eflatmajor += 1
            keys[str(full_key)] = eflatmajor
        if(str(full_key).lower() == "E Major".lower()):
            emajor += 1
            keys[str(full_key)] = emajor
        if(str(full_key).lower() == "F Major".lower()):
            fmajor += 1
            keys[str(full_key)] = fmajor
        if(str(full_key).lower() == "F# Major".lower()):
            fsharpmajor += 1
            keys[str(full_key)] = fsharpmajor
        if(str(full_key).lower() == "G Major".lower()):
            gmajor += 1
            keys[str(full_key)] = gmajor
        if(str(full_key).lower() == "Ab Major".lower()):
            aflatmajor += 1
            keys[str(full_key)] = aflatmajor
        if(str(full_key).lower() == "A Major".lower()):
            amajor += 1
            keys[str(full_key)] = amajor
        if(str(full_key).lower() == "Bb Major".lower()):
            bflatmajor += 1
            keys[str(full_key)] = bflatmajor
        if(str(full_key).lower() == "B Major".lower()):
            bmajor += 1
            keys[str(full_key)] = bmajor
        
        #minor keys
        if str(full_key) not in keys:                
                keys[str(full_key)] = 0
        if(str(full_key).lower() == "C Minor".lower()):
            cminor += 1
            keys[str(full_key)] = cminor
        if(str(full_key).lower() == "Db Minor".lower()):
            dflatminor += 1
            keys[str(full_key)] = dflatminor
        if(str(full_key).lower() == "D Minor".lower()):
            dminor += 1
            keys[str(full_key)] = dminor
        if(str(full_key).lower() == "Eb Minor".lower()):
            eflatminor += 1
            keys[str(full_key)] = eflatminor
        if(str(full_key).lower() == "E Minor".lower()):
            eminor += 1
            keys[str(full_key)] = eminor
        if(str(full_key).lower() == "F Minor".lower()):
            fminor += 1
            keys[str(full_key)] = fminor
        if(str(full_key).lower() == "F# Minor".lower()):
            fsharpminor += 1
            keys[str(full_key)] = fsharpminor
        if(str(full_key).lower() == "G Minor".lower()):
            gminor += 1
            keys[str(full_key)] = gminor
        if(str(full_key).lower() == "Ab Minor".lower()):
            aflatminor += 1
            keys[str(full_key)] = aflatminor
        if(str(full_key).lower() == "A Minor".lower()):
            aminor += 1
            keys[str(full_key)] = aminor
        if(str(full_key).lower() == "Bb Minor".lower()):
            bflatminor += 1
            keys[str(full_key)] = bflatminor
        if(str(full_key).lower() == "B Minor".lower()):
            bminor += 1
            keys[str(full_key)] = bminor
    return keys

def get_tempo(spreadsheet):
    allegro = 0
    vivace = 0
    andante = 0
    adagietto = 0
    moderato = 0
    allegmod = 0
    tempos = {}
    for j in range(11):
        cell = spreadsheet.cell(row = 6, column = j+1)
        if("Allegro Moderato" in cell.value):
            allegmod+=1
            tempos["Allegro Moderato"] = allegmod
        elif("Allegro" in cell.value):
            allegro+=1
            tempos["Allegro"] = allegro
        elif("Vivace" in cell.value):
            vivace+=1
            tempos["Vivace"] = vivace
        elif("Andante" in cell.value):
            andante+=1
            tempos["Andante"] = andante
        elif("Adagietto" in cell.value):
            adagietto+=1
            tempos["Adagietto"] = adagietto
        elif("Moderato" in cell.value):
            moderato+=1
            tempos["Moderato"] = moderato
    return tempos

def get_tempo_avg(spreadsheet):
    cell_nt = 0
    f = 2
    for j in range(10):
        cell = spreadsheet.cell(row = 9, column = f)
        cell_nt += int(cell.value)
        f+=1
    cell_nt /= 10
    return cell_nt

def get_chords(spreadsheet):
    chords_arr = []
    for j in range(11):
        cell = spreadsheet.cell(row = 5, column = j+1)
        chords_arr.append(cell.value)    
    del chords_arr[0]    
    return chords_arr

def get_chord_pairs(spreadsheet):
    chords = get_chords(spreadsheet)

    song0_1 = chords[0][0:3]
    song0_2 = chords[0][2:5]
    song0_3 = chords[0][4:7]
    song0 = [song0_1, song0_2, song0_3]

    song1_1 = chords[1][0:3]
    song1_2 = chords[1][2:5]
    song1_3 = chords[1][4:7]
    song1 = [song1_1, song1_2, song1_3]

    song2_1 = chords[2][0:3]
    song2_2 = chords[2][2:5]
    song2_3 = chords[2][4:7]
    song2 = [song2_1, song2_2, song2_3]
    
    song3_1 = chords[3][0:3]
    song3_2 = chords[3][2:5]
    song3_3 = chords[3][4:7]
    song3 = [song3_1, song3_2, song3_3]

    song4_1 = chords[4][0:3]
    song4_2 = chords[4][2:5]
    song4_3 = chords[4][4:7]
    song4 = [song4_1, song4_2, song4_3]

    song5_1 = chords[5][0:3]
    song5_2 = chords[5][2:5]
    song5_3 = chords[5][4:7]
    song5 = [song5_1, song5_2, song5_3]

    song6_1 = chords[6][0:3]
    song6_2 = chords[6][2:5]
    song6_3 = chords[6][4:7]
    song6 = [song6_1, song6_2, song6_3]

    song7_1 = chords[7][0:3]
    song7_2 = chords[7][2:5]
    song7_3 = chords[7][4:7]
    song7 = [song7_1, song7_2, song7_3]

    song8_1 = chords[8][0:3]
    song8_2 = chords[8][2:5]
    song8_3 = chords[8][4:7]
    song8 = [song8_1, song8_2, song8_3]

    song9_1 = chords[9][0:3]
    song9_2 = chords[9][2:5]
    song9_3 = chords[9][4:7]
    song9 = [song9_1, song9_2, song9_3]

    all_songs = song0 + song1 + song2 + song3 + song4 + song5 + song6 + song7 + song8 + song9
    return all_songs

def get_degrees(spreadsheet):
    chords = get_chords(spreadsheet)

    song0_1 = chords[0][0:1]
    song0_2 = chords[0][2:3]
    song0_3 = chords[0][4:5]
    song0_4 = chords[0][6:7]
    song0 = [song0_1, song0_2, song0_3, song0_4]

    song1_1 = chords[1][0:1]
    song1_2 = chords[1][2:3]
    song1_3 = chords[1][4:5]
    song1_4 = chords[1][6:7]
    song1 = [song1_1, song1_2, song1_3, song1_4]

    song2_1 = chords[2][0:1]
    song2_2 = chords[2][2:3]
    song2_3 = chords[2][4:5]
    song2_4 = chords[2][6:7]
    song2 = [song2_1, song2_2, song2_3, song2_4]
    
    song3_1 = chords[3][0:1]
    song3_2 = chords[3][2:3]
    song3_3 = chords[3][4:5]
    song3_4 = chords[3][6:7]
    song3 = [song3_1, song3_2, song3_3, song3_4]

    song4_1 = chords[4][0:1]
    song4_2 = chords[4][2:3]
    song4_3 = chords[4][4:5]
    song4_4 = chords[4][6:7]
    song4 = [song4_1, song4_2, song4_3, song4_4]

    song5_1 = chords[5][0:1]
    song5_2 = chords[5][2:3]
    song5_3 = chords[5][4:5]
    song5_4 = chords[5][6:7]
    song5 = [song5_1, song5_2, song5_3, song5_4]

    song6_1 = chords[6][0:1]
    song6_2 = chords[6][2:3]
    song6_3 = chords[6][4:5]
    song6_4 = chords[6][6:7]
    song6 = [song6_1, song6_2, song6_3, song6_4]

    song7_1 = chords[7][0:1]
    song7_2 = chords[7][2:3]
    song7_3 = chords[7][4:5]
    song7_4 = chords[7][6:7]
    song7 = [song7_1, song7_2, song7_3, song7_4]

    song8_1 = chords[8][0:1]
    song8_2 = chords[8][2:3]
    song8_3 = chords[8][4:5]
    song8_4 = chords[8][6:7]
    song8 = [song8_1, song8_2, song8_3, song8_4]

    song9_1 = chords[9][0:1]
    song9_2 = chords[9][2:3]
    song9_3 = chords[9][4:5]
    song9_4 = chords[9][6:7]
    song9 = [song9_1, song9_2, song9_3, song9_4]

    all_songs = song0 + song1 + song2 + song3 + song4 + song5 + song6 + song7 + song8 + song9
    return all_songs

def get_occurrences(spreadsheet):
    occur_dict = {}
    deg_arr = get_degrees(spreadsheet)
    for x in range(len(deg_arr)):
        if deg_arr[x] not in occur_dict:                
                occur_dict[deg_arr[x]] = 0
        occur_dict[deg_arr[x]] += 1

    return occur_dict

def analyze_chord_pairs(pair_arr):
    copyarr = pair_arr
    chord_pair_nums = {}
    
    for z in range(len(pair_arr)):
        for j in range(len(copyarr)):
            if pair_arr[j] not in chord_pair_nums:                
                chord_pair_nums[pair_arr[j]] = 0
            if(copyarr[j] == pair_arr[z]):
                chord_pair_nums[pair_arr[j]] += 1       
    for key in chord_pair_nums:
        chord_pair_nums[key] = int(math.sqrt(chord_pair_nums[key]))
    return chord_pair_nums


book = xlsxreader.load_workbook("bb10_data_for_analysis.xlsx")
charts_2011 = book["2011 Charts"]
charts_2012 = book["2012 Charts"]
charts_2013 = book["2013 Charts"]
charts_2014 = book["2014 Charts"]
charts_2015 = book["2015 Charts"]
charts_2016 = book["2016 Charts"]
charts_2017 = book["2017 Charts"]
charts_2018 = book["2018 Charts"]
charts_2019 = book["2019 Charts"]
charts_2020 = book["2020 Charts"]




#print("Tonics: " + str(collections.Counter(get_tempo(charts_2011)))
#print("Key Qualities: " + str(get_chord_pairs(charts_2011)))
#print("Tempos: " + str(get_tempo(charts_2011)))
#print("Tempo nums: " + str(get_tempo_nums(charts_2011)))
#print("Chords:" + str(get_chords(charts_2011)))
#print("Chord Pairs:" + str(get_chord_pairs(charts_2011)))
#print("Pairs:" + str(analyze_chord_pairs(get_chord_pairs(charts_2011))))
#print("degs: " + str(get_degrees(charts_2011)))

#find occurrences of certain chords relative to their scale degree of all songs in the dataset
occur_dict = dict(collections.Counter(get_occurrences(charts_2011)) + collections.Counter(get_occurrences(charts_2012)) + 
collections.Counter(get_occurrences(charts_2013)) + collections.Counter(get_occurrences(charts_2014)) + 
collections.Counter(get_occurrences(charts_2015)) + collections.Counter(get_occurrences(charts_2016)) + 
collections.Counter(get_occurrences(charts_2017)) + collections.Counter(get_occurrences(charts_2018)) +
collections.Counter(get_occurrences(charts_2019)) + collections.Counter(get_occurrences(charts_2020)))

#find tempos of all songs
all_tempos_dict = dict(collections.Counter(get_tempo(charts_2011)) + collections.Counter(get_tempo(charts_2012)) + 
collections.Counter(get_tempo(charts_2013)) + collections.Counter(get_tempo(charts_2014)) + 
collections.Counter(get_tempo(charts_2015)) + collections.Counter(get_tempo(charts_2016)) + 
collections.Counter(get_tempo(charts_2017)) + collections.Counter(get_tempo(charts_2018)) +
collections.Counter(get_tempo(charts_2019)) + collections.Counter(get_tempo(charts_2020)))
n = 0

#find average tempo of all songs in double variable
avg_tempo = (get_tempo_avg(charts_2011) + get_tempo_avg(charts_2012) + get_tempo_avg(charts_2013) + get_tempo_avg(charts_2014) + 
get_tempo_avg(charts_2015) + get_tempo_avg(charts_2016) + get_tempo_avg(charts_2017) + get_tempo_avg(charts_2018) + 
get_tempo_avg(charts_2019) + get_tempo_avg(charts_2020))
avg_tempo /= 10

all_tonics = dict(collections.Counter(get_tonic(charts_2011)) + collections.Counter(get_tonic(charts_2012)) + 
collections.Counter(get_tonic(charts_2013)) + collections.Counter(get_tonic(charts_2014)) + 
collections.Counter(get_tonic(charts_2015)) + collections.Counter(get_tonic(charts_2016)) + 
collections.Counter(get_tonic(charts_2017)) + collections.Counter(get_tonic(charts_2018)) + 
collections.Counter(get_tonic(charts_2019)) + collections.Counter(get_tonic(charts_2020)))

all_qualities = dict(collections.Counter(get_key_quality(charts_2011)) + collections.Counter(get_key_quality(charts_2012)) + 
collections.Counter(get_key_quality(charts_2013)) + collections.Counter(get_key_quality(charts_2014)) + 
collections.Counter(get_key_quality(charts_2015)) + collections.Counter(get_key_quality(charts_2016)) + 
collections.Counter(get_key_quality(charts_2017)) + collections.Counter(get_key_quality(charts_2018)) + 
collections.Counter(get_key_quality(charts_2019)) + collections.Counter(get_key_quality(charts_2020)))

all_chord_pairs = dict(collections.Counter(get_chord_pairs(charts_2011)) + collections.Counter(get_chord_pairs(charts_2012)) + 
collections.Counter(get_chord_pairs(charts_2013)) + collections.Counter(get_chord_pairs(charts_2014)) + 
collections.Counter(get_chord_pairs(charts_2015)) + collections.Counter(get_chord_pairs(charts_2016)) + 
collections.Counter(get_chord_pairs(charts_2017)) + collections.Counter(get_chord_pairs(charts_2018)) + 
collections.Counter(get_chord_pairs(charts_2019)) + collections.Counter(get_chord_pairs(charts_2020)))
all_chord_pairs_sort = sorted(all_chord_pairs.items(), key=operator.itemgetter(1)) 
allchords_sortdict = dict(all_chord_pairs_sort)

key_whole_dict = dict(collections.Counter(get_key_whole(charts_2011)) + collections.Counter(get_key_whole(charts_2012)) + 
collections.Counter(get_key_whole(charts_2013)) + collections.Counter(get_key_whole(charts_2014)) + 
collections.Counter(get_key_whole(charts_2015)) + collections.Counter(get_key_whole(charts_2016)) + 
collections.Counter(get_key_whole(charts_2017)) + collections.Counter(get_key_whole(charts_2018)) + 
collections.Counter(get_key_whole(charts_2019)) + collections.Counter(get_key_whole(charts_2020)))
key_whole_sort = sorted(key_whole_dict.items(), key=operator.itemgetter(1)) 
key_whole_sortdict = dict(key_whole_sort)

print("deg Occurrences: " + str(occur_dict))
print("tempo Occurrences: " + str(all_tempos_dict))
print("average tempo: " + str(avg_tempo))
print("Tonics: " + str(all_tonics))
print("Key Qualities: " + str(all_qualities))
print("Chord Pairs: " + str(allchords_sortdict))
print("Key Whole: " + str(key_whole_sortdict))
print("Tempo average 2011: " + str(get_tempo_avg(charts_2011)))
print("Tempo average 2012: " + str(get_tempo_avg(charts_2012)))
print("Tempo average 2013: " + str(get_tempo_avg(charts_2013)))
print("Tempo average 2014: " + str(get_tempo_avg(charts_2014)))
print("Tempo average 2015: " + str(get_tempo_avg(charts_2015)))
print("Tempo average 2016: " + str(get_tempo_avg(charts_2016)))
print("Tempo average 2017: " + str(get_tempo_avg(charts_2017)))
print("Tempo average 2018: " + str(get_tempo_avg(charts_2018)))
print("Tempo average 2019: " + str(get_tempo_avg(charts_2019)))
print("Tempo average 2020: " + str(get_tempo_avg(charts_2020)))

print("chord pairs 2011: " + str(dict(collections.Counter(get_chord_pairs(charts_2011)))))
print("chord pairs 2012: " + str(dict(collections.Counter(get_chord_pairs(charts_2012)))))
print("chord pairs 2013: " + str(dict(collections.Counter(get_chord_pairs(charts_2013)))))
print("chord pairs 2014: " + str(dict(collections.Counter(get_chord_pairs(charts_2014)))))
print("chord pairs 2015: " + str(dict(collections.Counter(get_chord_pairs(charts_2015)))))
print("chord pairs 2016: " + str(dict(collections.Counter(get_chord_pairs(charts_2016)))))
print("chord pairs 2017: " + str(dict(collections.Counter(get_chord_pairs(charts_2017)))))
print("chord pairs 2018: " + str(dict(collections.Counter(get_chord_pairs(charts_2018)))))
print("chord pairs 2019: " + str(dict(collections.Counter(get_chord_pairs(charts_2019)))))
print("chord pairs 2020: " + str(dict(collections.Counter(get_chord_pairs(charts_2020)))))

print("key qualities 2011: " + str(get_key_quality(charts_2011)))
print("key qualities 2012: " + str(get_key_quality(charts_2012)))
print("key qualities 2013: " + str(get_key_quality(charts_2013)))
print("key qualities 2014: " + str(get_key_quality(charts_2014)))
print("key qualities 2015: " + str(get_key_quality(charts_2015)))
print("key qualities 2016: " + str(get_key_quality(charts_2016)))
print("key qualities 2017: " + str(get_key_quality(charts_2017)))
print("key qualities 2018: " + str(get_key_quality(charts_2018)))
print("key qualities 2019: " + str(get_key_quality(charts_2019)))
print("key qualities 2020: " + str(get_key_quality(charts_2020)))
